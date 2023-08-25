from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
import re
import os

COL_NUM = 5
PATH = 'files/'
GEN_PATH = 'generated_files/'
LAST_WORD = ['Итого:', 'итого:', 'Общий итог:', 'общий итог', 'итого', 'общий итог', 'ИТОГО:', 'ИТОГО', 'Всего:']


def filename(path):
    pattern = re.compile(r"\W(\S+)")
    m = pattern.search(path)
    return m.group(1)  #filename


def merge_bold_center(sheet, row, col_start, col_end):
    sheet.merge_cells(start_row=row + 1, end_row=row + 1, start_column=col_start, end_column=col_end)
    cell = sheet.cell(row=row + 1, column=1)
    cell.alignment = Alignment(horizontal="center")
    cell.font = Font(bold=True)
    if row == 0:
        cell.font = Font(size="16")
    else:
        cell.value = str(cell.value).capitalize()
    return cell


def merge_bold_right(sheet, row, col_start, col_end):
    sheet.merge_cells(start_row=row + 1, end_row=row + 1, start_column=col_start, end_column=col_end)
    cell = sheet.cell(row=row + 1, column=1)
    cell.alignment = Alignment(horizontal="right")
    cell.font = Font(bold=True)
    return cell


def sheet_style(sheet, length):
    sheet.column_dimensions['A'].width = length
    sheet.column_dimensions['B'].alignment = Alignment(horizontal="center")


def write_file(data, path):
    '''
    :param data: list of values
    :param path: path to source file to create a new file with the same name
    :return: creates a new file
    '''
    book = Workbook()
    sheet = book.active

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    for row in range(len(data)):
        for col in range(COL_NUM):
            cell = sheet.cell(row=row+1, column=col+1)
            cell.value = data[row][col]
            cell.border = border

            # Set the alignment for columns
            if col in [1, 2, 3]:
                cell.alignment = Alignment(horizontal='center')
            if col == 4:
                cell.alignment = Alignment(horizontal='right')
            if row == 1:
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)

        # Style the sub-title
        if data[row][1:] == [None, None, None, None]:
            merge_bold_center(sheet, row, 1, COL_NUM)

        # Insert formula for total price
        # Style the last row
        if data[row][0] in LAST_WORD:
            merge_bold_right(sheet, row, 1, COL_NUM-1)
            cell = sheet.cell(row=row + 1, column=COL_NUM)
            cell.value = f'=SUM(E1:E{row})'
            cell.font = Font(bold=True)

        # Insert Formulas
        # data[row][2] - unit_cost
        # data[row][3] - quantity
        if data[row][2] is not None and data[row][3] is not None and not isinstance(data[row][2], str):
            cell = sheet.cell(row=row + 1, column=COL_NUM)
            cell.value = f'=C{row + 1}*D{row + 1}'

    # Find the longest word in the first column
    # Set the width of the first column
    length = max(len(cell[:][0]) for cell in data[1:]) + 1
    sheet_style(sheet, length)
    if not os.path.exists(GEN_PATH):
        os.makedirs(GEN_PATH)
    book.save(GEN_PATH + filename(path))


def read_file(sheet, columns=COL_NUM):
    '''
    turning the sheet into list, without empty lines, without formulas
    prints out the math error in the file
    :param sheet: list of xls file
    :param columns: number of valuable columns (5 - default)
    :return: list of values
    '''
    data = []
    is_last_row = False  # last line flag
    row_num = 1
    total = 0
    while True:
        row = []

        # Read the row, remove spaces
        for col_num in range(1, columns+1):
            val = sheet.cell(row=row_num, column=col_num).value
            if ',' in str(val) and col_num != (1 or 2):
                val = float(str(val).replace(',', '.'))
            if col_num == 2 and isinstance(val, str):
                val = "".join(val.split())
                val = val.replace('.', '')
            if isinstance(val, str):
                val = val.strip()
            if col_num == 1 and val is not None:
                val = str(val).capitalize()
            if val in LAST_WORD:
                is_last_row = True
            row.append(val)

        # Skip empty lines
        # Leave empty line for the name if there is no name
        if row == [None, None, None, None, None] and row_num != 1:
            row_num += 1
            continue

        # Count the price for current work
        # Compare the result with actual value
        unit_cost = row[columns-3]
        quantity = row[columns-2]
        total_price = row[columns-1]

        if ('=' in str(total_price) or
            ((quantity is not None) and not isinstance(unit_cost, str))) and \
                unit_cost is not None:
            result = int(quantity * unit_cost)
            if '=' not in str(total_price) and result != total_price:
                print(f'{row_num} == {row} --> {result} ')
            row[columns-1] = result

        # Total value for the last row in the file
        if is_last_row:
            if row[columns-1] != total and '=' not in str(row[columns-1]):
                print(f'total == {row} --> {total} ')
            row[columns-1] = total
            data.append(row)
            break
        else:
            if isinstance(row[columns - 1], int):
                total += row[columns - 1]
            data.append(row)

        row_num += 1
    return data


def load_file(path):
    '''
    loading xls file from path, writes modified data to a file
    :param path: path to xlsx file
    '''
    book = load_workbook(path)
    sheet = book.active
    data = read_file(sheet)
    write_file(data, path)


if __name__ == "__main__":
    file_name = 'smeta_05_17_2023.xlsx'
    print(PATH + file_name)
    load_file(PATH + file_name)
    # for file_name in os.listdir('files/'):
    #     print(PATH + file_name)
    #     load_file(PATH + file_name)


